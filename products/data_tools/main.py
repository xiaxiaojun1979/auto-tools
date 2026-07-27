#!/usr/bin/env python3
"""
数据清洗工具包 (Data Cleaning Toolkit v1.0)
Author: AutoTools Studio
自动清理 CSV/Excel 数据：去重、格式化、标准化、校验
"""

import os
import sys
import csv
import json
import re
from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict
import argparse
from concurrent.futures import ProcessPoolExecutor
import hashlib

try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False


class DataCleaner:
    """数据清洗引擎"""

    def __init__(self, verbose=True):
        self.verbose = verbose
        self.stats = {"总行数": 0, "重复行": 0, "空值行": 0, "清洗后行数": 0, "错误": 0}

    def log(self, msg):
        if self.verbose:
            print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

    # ========== 读取 ==========
    def read_file(self, filepath):
        """读取 CSV 或 Excel"""
        path = Path(filepath)
        if not path.exists():
            print(f"[X] 文件不存在: {filepath}")
            return None, None

        if path.suffix.lower() in ('.csv', '.tsv'):
            delimiter = '\t' if path.suffix.lower() == '.tsv' else ','
            return self._read_csv(path, delimiter)
        elif path.suffix.lower() in ('.xlsx', '.xls'):
            return self._read_excel(path)
        else:
            print(f"[X] 不支持的文件格式: {path.suffix}")
            return None, None

    def _read_csv(self, path, delimiter):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
            if not rows:
                return None, None
            # 尝试检测编码
            return rows[0], rows[1:]  # header, data
        except UnicodeDecodeError:
            # 尝试其他编码
            for enc in ['gbk', 'gb2312', 'utf-8-sig']:
                try:
                    with open(path, 'r', encoding=enc) as f:
                        reader = csv.reader(f)
                        rows = list(reader)
                    if rows:
                        return rows[0], rows[1:]
                except:
                    continue
            print(f"[X] 无法解码文件: {path}")
            return None, None

    def _read_excel(self, path):
        if not HAS_EXCEL:
            print("[X] 需要安装 openpyxl: pip3 install openpyxl")
            return None, None
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, None
        header = [str(c) if c else f"列{i}" for i, c in enumerate(rows[0])]
        data = [list(r) for r in rows[1:]]
        return header, data

    # ========== 功能1: 去除重复行 ==========
    def remove_duplicates(self, rows, key_columns=None, keep='first'):
        """去除重复行"""
        seen = set()
        unique = []
        dup_count = 0

        for row in rows:
            if key_columns:
                key = tuple(str(row[i]) if i < len(row) else '' for i in key_columns)
            else:
                key = tuple(str(c) for c in row)

            if key in seen:
                dup_count += 1
                continue
            seen.add(key)
            unique.append(row)

        self.stats["重复行"] += dup_count
        self.log(f"去除重复行: {dup_count} 行")
        return unique

    # ========== 功能2: 清理空值 ==========
    def clean_empty(self, rows, strategy='drop_row', fill_value=''):
        """处理空值"""
        cleaned = []
        empty_count = 0

        for row in rows:
            empty_cols = [i for i, c in enumerate(row) if c is None or str(c).strip() == '']

            if strategy == 'drop_row':
                if empty_cols:
                    empty_count += 1
                    continue
                cleaned.append(row)

            elif strategy == 'fill_empty':
                new_row = list(row)
                for i in empty_cols:
                    new_row[i] = fill_value
                cleaned.append(new_row)
                empty_count += len(empty_cols)

            elif strategy == 'drop_col':
                # 标记删除
                new_row = [c for i, c in enumerate(row) if i not in empty_cols]
                cleaned.append(new_row)

        self.stats["空值行"] += empty_count if strategy == 'drop_row' else 0
        self.log(f"处理空值: {empty_count} 处")
        return cleaned

    # ========== 功能3: 格式化数据 ==========
    def format_column(self, rows, column, fmt_type):
        """格式化指定列"""
        formatted = []
        for row in rows:
            new_row = list(row)
            if column < len(row):
                val = str(row[column]).strip()
                try:
                    if fmt_type == 'lower':
                        new_row[column] = val.lower()
                    elif fmt_type == 'upper':
                        new_row[column] = val.upper()
                    elif fmt_type == 'title':
                        new_row[column] = val.title()
                    elif fmt_type == 'strip':
                        new_row[column] = val.strip()
                    elif fmt_type == 'phone':
                        # 手机号格式化
                        digits = re.sub(r'\D', '', val)
                        if len(digits) == 11:
                            new_row[column] = f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
                    elif fmt_type == 'date':
                        # 日期标准化
                        new_row[column] = self._normalize_date(val)
                    elif fmt_type == 'number':
                        new_row[column] = self._normalize_number(val)
                except:
                    self.stats["错误"] += 1
            formatted.append(new_row)

        self.log(f"格式化列 {column} -> {fmt_type}")
        return formatted

    def _normalize_date(self, val):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y年%m月%d日',
                     '%m/%d/%Y', '%d/%m/%Y', '%Y%m%d']:
            try:
                return datetime.strptime(val, fmt).strftime('%Y-%m-%d')
            except:
                continue
        return val

    def _normalize_number(self, val):
        # 去除货币符号和逗号
        val = re.sub(r'[¥$€￥,，\s]', '', val)
        try:
            return f"{float(val):.2f}"
        except:
            return val

    # ========== 功能4: 数据校验 ==========
    def validate_data(self, rows, rules=None):
        """数据校验"""
        if rules is None:
            rules = []

        errors = []
        col_count = len(rows[0]) if rows else 0

        builtin_rules = {
            'not_empty': lambda v: v is not None and str(v).strip() != '',
            'is_number': lambda v: re.match(r'^-?\d+(\.\d+)?$', str(v).strip()) is not None,
            'is_email': lambda v: re.match(r'^[\w.+-]+@[\w-]+\.[\w.]+$', str(v).strip()) is not None,
            'is_phone': lambda v: re.sub(r'\D', '', str(v)).isdigit() and len(re.sub(r'\D', '', str(v))) == 11,
            'is_url': lambda v: str(v).strip().startswith('http'),
        }

        for i, row in enumerate(rows):
            for col_idx, rule in rules:
                if col_idx < len(row):
                    val = row[col_idx]
                    if rule in builtin_rules and not builtin_rules[rule](val):
                        errors.append((i, col_idx, rule, val))
                        self.stats["错误"] += 1

        if errors:
            print(f"\n⚠️  发现 {len(errors)} 个数据问题:")
            for row_idx, col_idx, rule, val in errors[:20]:
                print(f"  行{row_idx+1} 列{col_idx+1}: [{rule}] 值='{val}'")
            if len(errors) > 20:
                print(f"  ... 还有 {len(errors)-20} 个问题")

        return errors

    # ========== 功能5: 统计分析 ==========
    def analyze(self, rows, header):
        """数据统计分析"""
        print(f"\n{'='*50}")
        print(f"  数据统计分析报告")
        print(f"{'='*50}")
        print(f"  总行数: {len(rows)}")
        print(f"  总列数: {len(header) if header else 0}")
        print()

        for i, col in enumerate(header):
            values = [str(r[i]) for r in rows if i < len(r) and r[i] is not None]
            if not values:
                continue

            unique_vals = len(set(values))
            empty_count = sum(1 for v in values if v.strip() == '')
            val_lengths = [len(v) for v in values]

            print(f"  📊 {col}:")
            print(f"      非空值: {len(values)} | 唯一值: {unique_vals} | 空值: {empty_count}")
            if val_lengths:
                print(f"      最短: {min(val_lengths)}字 | 最长: {max(val_lengths)}字 | 平均: {sum(val_lengths)//len(val_lengths)}字")

            # 显示最常见值
            if unique_vals <= 10:
                counter = Counter(values)
                print(f"      值分布: {dict(counter.most_common(5))}")

        print(f"{'='*50}")
        return self.stats

    # ========== 功能6: 导出 ==========
    def export_csv(self, rows, header, filepath):
        """导出为CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            if header:
                w.writerow(header)
            w.writerows(rows)
        print(f"[✓] 已导出 {len(rows)} 行到 {filepath}")
        return filepath

    def export_excel(self, rows, header, filepath, sheet_name="清洗结果"):
        """导出为Excel"""
        if not HAS_EXCEL:
            print("[X] 需要安装 openpyxl")
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if header:
            ws.append(header)
        for row in rows:
            ws.append(row)

        wb.save(filepath)
        print(f"[✓] 已导出 {len(rows)} 行到 {filepath}")
        return filepath


def main():
    parser = argparse.ArgumentParser(
        description="🧹 数据清洗工具包 v1.0 - 一键清洗你的数据",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 查看文件统计
  python3 main.py analyze data.csv

  # 去除重复行
  python3 main.py clean data.csv --dedup

  # 去除重复行（基于特定列）
  python3 main.py clean data.csv --dedup --keys 0 1

  # 去除空值行
  python3 main.py clean data.csv --drop-empty

  # 格式化列
  python3 main.py clean data.csv --format-col 0 upper --format-col 2 phone

  # 数据校验
  python3 main.py validate data.csv --rule 0 not_empty --rule 1 is_email

  # 完整流水线
  python3 main.py pipeline data.csv --dedup --drop-empty --format-col 0 title --output cleaned.csv
        """
    )
    subparsers = parser.add_subparsers(dest="command")

    # analyze
    p_ana = subparsers.add_parser("analyze", help="数据分析")
    p_ana.add_argument("file", help="数据文件")

    # clean
    p_clean = subparsers.add_parser("clean", help="清洗数据")
    p_clean.add_argument("file", help="数据文件")
    p_clean.add_argument("--dedup", action="store_true", help="去除重复行")
    p_clean.add_argument("--keys", nargs="*", type=int, help="去重依据列索引")
    p_clean.add_argument("--drop-empty", action="store_true", help="去除空值行")
    p_clean.add_argument("--fill-empty", default="", help="填充空值")
    p_clean.add_argument("--format-col", nargs=2, action="append",
                         metavar=("COL", "TYPE"), help="格式化列")
    p_clean.add_argument("--output", default="", help="输出文件")

    # validate
    p_val = subparsers.add_parser("validate", help="数据校验")
    p_val.add_argument("file", help="数据文件")
    p_val.add_argument("--rule", nargs=2, action="append",
                       metavar=("COL", "RULE"), required=True, help="校验规则")
    p_val.add_argument("--output", default="", help="输出校验报告")

    # pipeline
    p_pipe = subparsers.add_parser("pipeline", help="完整清洗流水线")
    p_pipe.add_argument("file", help="数据文件")
    p_pipe.add_argument("--dedup", action="store_true", help="去重")
    p_pipe.add_argument("--keys", nargs="*", type=int, help="去重列")
    p_pipe.add_argument("--drop-empty", action="store_true", help="去空值")
    p_pipe.add_argument("--fill-empty", default="", help="填充空值")
    p_pipe.add_argument("--format-col", nargs=2, action="append", help="格式化列")
    p_pipe.add_argument("--output", default="", help="输出文件")
    p_pipe.add_argument("--validate", action="store_true", help="校验后输出报告")

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        return

    cleaner = DataCleaner()
    header, data = cleaner.read_file(args.file)
    if header is None:
        return

    print(f"[✓] 读取完成: {len(data)} 行, {len(header)} 列")
    print(f"    列名: {header}")

    if args.command == "analyze":
        cleaner.analyze(data, header)
        return

    if args.command in ("clean", "pipeline"):
        if args.dedup:
            data = cleaner.remove_duplicates(data, args.keys)

        if args.drop_empty:
            data = cleaner.clean_empty(data, 'drop_row')

        if args.fill_empty:
            data = cleaner.clean_empty(data, 'fill_empty', args.fill_empty)

        if args.format_col:
            for col, fmt in args.format_col:
                data = cleaner.format_column(data, int(col), fmt)

        out_path = args.output or f"cleaned_{Path(args.file).name}"
        cleaner.export_csv(data, header, out_path)

        cleaner.stats["总行数"] = len(data) + (cleaner.stats.get("重复行", 0))
        cleaner.stats["清洗后行数"] = len(data)
        print(f"\n📊 清洗总结: 原始约{cleaner.stats['总行数']}行 → 清洗后{cleaner.stats['清洗后行数']}行")

        if args.command == "pipeline" and args.validate:
            cleaner.validate_data(data, args.rules)

    elif args.command == "validate":
        rules = [(int(c), r) for c, r in (args.rule or [])]
        errors = cleaner.validate_data(data, rules)
        print(f"\n共发现 {len(errors)} 个数据问题")


if __name__ == "__main__":
    main()
